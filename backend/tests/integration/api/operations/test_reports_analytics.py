"""Phase 14 reports and analytics coverage.

Integration tests run against the Docker-backed backend + postgres + redis stack.
Seed data must be loaded before execution (make seed).

Run:
  python -m pytest tests/integration/api/operations/test_reports_analytics.py -v
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import uuid
from datetime import date, datetime, time, timedelta, timezone

import httpx
import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.core.database import async_session
from app.models.billing import Invoice
from app.models.erp import AttendanceRecord, AttendanceSession
from app.models.iam import Session
from app.models.lms import Assignment, Course, Grade, Submission
from app.models.reporting import DataExport

SCHOOL_ID = "00000000-0000-4000-8000-000000000001"


def _auth_headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


async def _get_first_class_id(client: httpx.AsyncClient, token: str) -> str:
    response = await client.get(
        "/reports/options",
        headers=_auth_headers(token),
        params={"type": "class_summary"},
    )
    assert response.status_code == 200
    classes = response.json()["data"]["classes"]
    assert classes, "Expected at least one reportable class in seed data"
    return classes[0]["id"]


async def _wait_for_report_ready(
    client: httpx.AsyncClient,
    token: str,
    job_id: str,
    *,
    timeout_seconds: int = 30,
) -> tuple[dict, float]:
    started = asyncio.get_running_loop().time()

    while True:
        elapsed = asyncio.get_running_loop().time() - started
        if elapsed > timeout_seconds:
            pytest.fail(
                f"Report {job_id} did not become ready within {timeout_seconds}s"
            )

        response = await client.get(
            f"/reports/{job_id}/status",
            headers=_auth_headers(token),
        )
        assert response.status_code == 200
        payload = response.json()["data"]
        if payload["status"] == "ready":
            return payload, elapsed
        if payload["status"] == "failed":
            pytest.fail(payload.get("error_message") or f"Report {job_id} failed")
        await asyncio.sleep(2)


def _overview_metric(response_json: dict, key: str) -> float:
    metrics = {
        item["key"]: item["value"]["current"]
        for item in response_json["data"]["metrics"]
    }
    return float(metrics[key])


def _date_bounds(from_date: date, to_date: date) -> tuple[datetime, datetime]:
    start_dt = datetime.combine(from_date, time.min, tzinfo=timezone.utc)
    end_dt = datetime.combine(
        to_date + timedelta(days=1), time.min, tzinfo=timezone.utc
    )
    return start_dt, end_dt


class TestReportsAndAnalyticsIntegration:
    @pytest.mark.asyncio
    async def test_student_report_card_download_and_cache_hit(
        self,
        client: httpx.AsyncClient,
        student_token: str,
    ):
        generate_response = await client.post(
            "/reports/generate",
            headers=_auth_headers(student_token),
            json={
                "type": "student_report_card",
                "locale": "fr",
                "from_date": "2020-01-01",
                "to_date": date.today().isoformat(),
            },
        )
        assert generate_response.status_code == 200
        job = generate_response.json()["data"]

        ready_job, _elapsed = await _wait_for_report_ready(
            client,
            student_token,
            job["id"],
        )
        assert ready_job["download_url"]
        token_download_response = await client.get(
            ready_job["download_url"]
            .replace("http://localhost:8000/api/v1", "")
            .removeprefix("/api/v1")
        )
        assert token_download_response.status_code == 200
        assert token_download_response.headers["content-type"].startswith(
            "application/pdf"
        )
        assert token_download_response.content.startswith(b"%PDF")

        download_response = await client.get(
            f"/reports/{job['id']}/download",
            headers=_auth_headers(student_token),
        )
        assert download_response.status_code == 200
        assert download_response.headers["content-type"].startswith("application/pdf")
        assert download_response.content.startswith(b"%PDF")

        cached_response = await client.post(
            "/reports/generate",
            headers=_auth_headers(student_token),
            json={
                "type": "student_report_card",
                "locale": "fr",
                "from_date": "2020-01-01",
                "to_date": date.today().isoformat(),
            },
        )
        assert cached_response.status_code == 200
        cached_job = cached_response.json()["data"]
        assert cached_job["cache_hit"] is True
        assert cached_job["id"] == ready_job["id"]

    @pytest.mark.asyncio
    async def test_arabic_class_summary_pdf_ready_within_30_seconds(
        self,
        client: httpx.AsyncClient,
        admin_token: str,
    ):
        class_id = await _get_first_class_id(client, admin_token)

        generate_response = await client.post(
            "/reports/generate",
            headers=_auth_headers(admin_token),
            json={
                "type": "class_summary",
                "class_id": class_id,
                "locale": "ar",
                "from_date": "2020-01-01",
                "to_date": date.today().isoformat(),
            },
        )
        assert generate_response.status_code == 200
        job = generate_response.json()["data"]

        ready_job, elapsed = await _wait_for_report_ready(
            client,
            admin_token,
            job["id"],
        )
        assert elapsed <= 30
        assert ready_job["parameters"]["locale"] == "ar"

        download_response = await client.get(
            f"/reports/{job['id']}/download",
            headers=_auth_headers(admin_token),
        )
        assert download_response.status_code == 200
        assert download_response.content.startswith(b"%PDF")

    @pytest.mark.asyncio
    async def test_csv_and_xlsx_exports_capture_audit_rows_and_fields(
        self,
        client: httpx.AsyncClient,
        admin_token: str,
    ):
        filters = {"from_date": "2020-01-01", "to_date": date.today().isoformat()}

        async with async_session() as session:
            before_count = await session.scalar(select(func.count(DataExport.id)))

        csv_response = await client.get(
            "/export/csv",
            headers=_auth_headers(admin_token),
            params={"entity": "grades", "filters": json.dumps(filters)},
        )
        assert csv_response.status_code == 200
        csv_rows = list(csv.DictReader(io.StringIO(csv_response.text)))
        assert csv_rows
        assert {
            "student_id",
            "subject",
            "assignment_title",
            "score",
            "total_points",
            "feedback",
            "published_at",
        }.issubset(csv_rows[0].keys())

        xlsx_response = await client.get(
            "/export/xlsx",
            headers=_auth_headers(admin_token),
            params={"entity": "grades", "filters": json.dumps(filters)},
        )
        assert xlsx_response.status_code == 200
        workbook = load_workbook(io.BytesIO(xlsx_response.content), read_only=True)
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        assert list(header_row) == [
            "student_id",
            "subject",
            "assignment_title",
            "score",
            "total_points",
            "feedback",
            "published_at",
        ]

        async with async_session() as session:
            after_count = await session.scalar(select(func.count(DataExport.id)))
            latest_logs = (
                (
                    await session.execute(
                        select(DataExport)
                        .order_by(DataExport.created_at.desc())
                        .limit(2)
                    )
                )
                .scalars()
                .all()
            )

        assert after_count == before_count + 2
        assert {log.format for log in latest_logs} == {"csv", "xlsx"}
        assert {log.entity for log in latest_logs} == {"grades"}

    @pytest.mark.asyncio
    async def test_analytics_overview_matches_raw_sql(
        self,
        client: httpx.AsyncClient,
        admin_token: str,
    ):
        from_date = date(2020, 1, 1)
        to_date = date.today()
        from_dt, to_dt = _date_bounds(from_date, to_date)

        response = await client.get(
            "/analytics/overview",
            headers=_auth_headers(admin_token),
            params={
                "from": from_date.isoformat(),
                "to": to_date.isoformat(),
                "compare": "false",
            },
        )
        assert response.status_code == 200

        async with async_session() as session:
            active_users = await session.scalar(
                select(func.count(func.distinct(Session.user_id))).where(
                    Session.school_id == uuid.UUID(SCHOOL_ID),
                    Session.created_at >= from_dt,
                    Session.created_at <= to_dt,
                )
            )

            attendance_result = await session.execute(
                select(
                    func.count(AttendanceRecord.id),
                    func.count().filter(AttendanceRecord.status == "present"),
                )
                .select_from(AttendanceRecord)
                .join(
                    AttendanceSession,
                    AttendanceSession.id == AttendanceRecord.attendance_session_id,
                )
                .where(
                    AttendanceRecord.school_id == uuid.UUID(SCHOOL_ID),
                    AttendanceSession.session_date >= from_date,
                    AttendanceSession.session_date <= to_date,
                )
            )
            total_attendance, present_attendance = attendance_result.one()
            attendance_rate = (
                (float(present_attendance or 0) / float(total_attendance)) * 100
                if total_attendance
                else 0.0
            )

            average_grade = await session.scalar(
                select(func.avg(Grade.score))
                .select_from(Grade)
                .join(Submission, Submission.id == Grade.submission_id)
                .join(Assignment, Assignment.id == Submission.assignment_id)
                .join(Course, Course.id == Assignment.course_id)
                .where(
                    Course.school_id == uuid.UUID(SCHOOL_ID),
                    Grade.created_at >= from_dt,
                    Grade.created_at <= to_dt,
                )
            )

            invoiced = await session.scalar(
                select(func.sum(Invoice.total_amount)).where(
                    Invoice.school_id == uuid.UUID(SCHOOL_ID),
                    Invoice.issued_date >= from_date,
                    Invoice.issued_date <= to_date,
                )
            )
            paid = await session.scalar(
                select(func.sum(Invoice.total_amount)).where(
                    Invoice.school_id == uuid.UUID(SCHOOL_ID),
                    Invoice.status == "paid",
                    Invoice.issued_date >= from_date,
                    Invoice.issued_date <= to_date,
                )
            )
            collection_rate = (
                (float(paid or 0) / float(invoiced)) * 100 if invoiced else 0.0
            )

        assert _overview_metric(response.json(), "active_users") == pytest.approx(
            float(active_users or 0),
            abs=0.01,
        )
        assert _overview_metric(response.json(), "attendance_rate") == pytest.approx(
            attendance_rate,
            abs=0.01,
        )
        assert _overview_metric(response.json(), "average_grade") == pytest.approx(
            float(average_grade or 0),
            abs=0.01,
        )
        assert _overview_metric(response.json(), "collection_rate") == pytest.approx(
            collection_rate,
            abs=0.01,
        )

    @pytest.mark.asyncio
    async def test_rbac_and_deny_ordering_for_reports_and_analytics(
        self,
        client: httpx.AsyncClient,
        admin_token: str,
        student_token: str,
    ):
        unauth_reports = await client.get("/reports")
        assert unauth_reports.status_code == 401

        unauth_analytics = await client.get("/analytics/overview")
        assert unauth_analytics.status_code == 401

        class_id = await _get_first_class_id(client, admin_token)
        forbidden_generate = await client.post(
            "/reports/generate",
            headers=_auth_headers(student_token),
            json={
                "type": "class_summary",
                "class_id": class_id,
                "locale": "fr",
            },
        )
        assert forbidden_generate.status_code == 403

        not_found_status = await client.get(
            f"/reports/{uuid.uuid4()}/status",
            headers=_auth_headers(student_token),
        )
        assert not_found_status.status_code == 404

        forbidden_analytics = await client.get(
            "/analytics/overview",
            headers=_auth_headers(student_token),
        )
        assert forbidden_analytics.status_code == 403
